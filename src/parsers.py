from __future__ import annotations

import re
from io import BytesIO

# Pattern for a single t.me link — used by normalize_identifier.
# Captures the username from: https://t.me/user, t.me/user, t.me/+invite, t.me/user/123
_TME_LINK_RE = re.compile(
    r"^(?:https?://)?t\.me/\+?([a-zA-Z][a-zA-Z0-9_]{3,31})(?:/\d+)?$"
)

# Bare username pattern (without @) — 4-32 chars, starts with letter.
_BARE_USERNAME_RE = re.compile(r"^[a-zA-Z][a-zA-Z0-9_]{3,31}$")


def normalize_identifier(raw: str) -> tuple[str, str]:
    """Normalize any Telegram identifier to ``(clean_value, kind)``.

    Accepts:
    - ``@username`` / ``username``
    - ``t.me/username`` / ``https://t.me/username`` / ``https://t.me/username/123``
    - Numeric IDs: ``-100123456789``, ``123456``

    Returns ``(clean_value, kind)`` where *kind* is:
    - ``'username'`` — *clean_value* is the lowercase username without ``@``
    - ``'numeric_id'`` — *clean_value* is the numeric string (may be negative)
    - ``'unknown'`` — couldn't parse; *clean_value* is the stripped original
    """
    raw = raw.strip()
    if not raw:
        return raw, "unknown"

    # t.me link (with or without scheme, with optional post id)
    m = _TME_LINK_RE.match(raw)
    if m:
        return m.group(1).lower(), "username"

    # @username
    if raw.startswith("@"):
        candidate = raw[1:]
        if _BARE_USERNAME_RE.match(candidate):
            return candidate.lower(), "username"

    # Numeric ID (positive or negative)
    if raw.lstrip("-").isdigit():
        return raw, "numeric_id"

    # Bare username
    if _BARE_USERNAME_RE.match(raw):
        return raw.lower(), "username"

    return raw, "unknown"


# Compiled pattern for extracting Telegram identifiers from arbitrary text.
# Order matters: full URLs first, then bare t.me/, then @username, then negative IDs.
_IDENTIFIER_RE = re.compile(
    r"https?://t\.me/[^\s\"'<>,;)]+"  # full t.me link
    r"|(?<![a-zA-Z0-9/])t\.me/[^\s\"'<>,;)]+"  # bare t.me/ (negative lookbehind)
    r"|@[a-zA-Z][a-zA-Z0-9_]{3,31}"  # @username (4-32 chars total)
    r"|-1\d{9,}",  # negative Telegram ID (-100...)
)


def parse_identifiers(text: str) -> list[str]:
    """Split text into channel identifiers.

    Supports separators: newline, comma, semicolon, tab.
    Strips whitespace and quotes. Skips empty lines and comments (#).
    """
    results: list[str] = []
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        for sep in (",", ";", "\t"):
            line = line.replace(sep, "\n")
        for part in line.splitlines():
            part = part.strip().strip("\"'").strip()
            if part:
                results.append(part)
    return results


def extract_identifiers(text: str) -> list[str]:
    """Extract Telegram identifiers from arbitrary text via regex.

    Finds t.me links, @usernames, and negative numeric IDs in any surrounding text.
    """
    return _IDENTIFIER_RE.findall(text)


def parse_file(content: bytes, filename: str = "") -> list[str]:
    """Universal file parser: extract Telegram identifiers from any file bytes.

    Detects xlsx by magic bytes (PK zip header), otherwise treats as text.
    """
    # Detect xlsx by magic bytes
    if content[:4] == b"PK\x03\x04":
        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                for cell_value in row:
                    if cell_value is not None:
                        parts.append(str(cell_value))
        wb.close()
        text = "\n".join(parts)
    else:
        # Text file: try utf-8-sig first, fallback to latin-1
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")

    return extract_identifiers(text)


def deduplicate_identifiers(identifiers: list[str]) -> list[str]:
    """Deduplicate identifiers case-insensitively, preserving order."""
    seen: set[str] = set()
    result: list[str] = []
    for ident in identifiers:
        key = ident.lower().strip()
        if key and key not in seen:
            seen.add(key)
            result.append(ident)
    return result
