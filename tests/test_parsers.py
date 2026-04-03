from __future__ import annotations

from src.parsers import (
    deduplicate_identifiers,
    extract_identifiers,
    normalize_identifier,
    parse_file,
    parse_identifiers,
)


class TestParseIdentifiers:
    def test_newlines(self):
        text = "@channel1\n@channel2\n@channel3"
        assert parse_identifiers(text) == ["@channel1", "@channel2", "@channel3"]

    def test_commas(self):
        text = "@ch1, @ch2, @ch3"
        assert parse_identifiers(text) == ["@ch1", "@ch2", "@ch3"]

    def test_semicolons(self):
        text = "@ch1; @ch2; @ch3"
        assert parse_identifiers(text) == ["@ch1", "@ch2", "@ch3"]

    def test_tabs(self):
        text = "@ch1\t@ch2\t@ch3"
        assert parse_identifiers(text) == ["@ch1", "@ch2", "@ch3"]

    def test_mixed_separators(self):
        text = "@ch1, @ch2\n@ch3; @ch4"
        assert parse_identifiers(text) == ["@ch1", "@ch2", "@ch3", "@ch4"]

    def test_tme_links(self):
        text = "t.me/channel1\nhttps://t.me/channel2"
        result = parse_identifiers(text)
        assert result == ["t.me/channel1", "https://t.me/channel2"]

    def test_numeric_ids(self):
        text = "-1001234567890\n-1009876543210"
        result = parse_identifiers(text)
        assert result == ["-1001234567890", "-1009876543210"]

    def test_empty_lines_skipped(self):
        text = "@ch1\n\n\n@ch2\n  \n@ch3"
        assert parse_identifiers(text) == ["@ch1", "@ch2", "@ch3"]

    def test_comments_skipped(self):
        text = "# This is a comment\n@ch1\n# Another comment\n@ch2"
        assert parse_identifiers(text) == ["@ch1", "@ch2"]

    def test_quotes_stripped(self):
        text = "\"@ch1\", '@ch2'"
        assert parse_identifiers(text) == ["@ch1", "@ch2"]

    def test_empty_input(self):
        assert parse_identifiers("") == []
        assert parse_identifiers("  \n  \n  ") == []


class TestExtractIdentifiers:
    def test_full_tme_links(self):
        text = "Check https://t.me/channel1 and https://t.me/channel2 for info"
        result = extract_identifiers(text)
        assert "https://t.me/channel1" in result
        assert "https://t.me/channel2" in result

    def test_bare_tme_links(self):
        text = "Go to t.me/mychannel for updates"
        result = extract_identifiers(text)
        assert "t.me/mychannel" in result

    def test_usernames(self):
        text = "Follow @testchannel and @another_one for news"
        result = extract_identifiers(text)
        assert "@testchannel" in result
        assert "@another_one" in result

    def test_negative_ids(self):
        text = "Channel ID: -1001234567890, also -1009876543210"
        result = extract_identifiers(text)
        assert "-1001234567890" in result
        assert "-1009876543210" in result

    def test_mixed_in_garbage(self):
        text = """
        Some random text here. Visit https://t.me/news_channel for breaking news.
        Also check @crypto_signals — it's great! The ID is -1001122334455.
        More junk: t.me/another_one is also good.
        Not a channel: example.com, hello@world.com
        """
        result = extract_identifiers(text)
        assert "https://t.me/news_channel" in result
        assert "@crypto_signals" in result
        assert "-1001122334455" in result
        assert "t.me/another_one" in result

    def test_no_false_positives_on_email(self):
        text = "email user@example.com shouldn't match short @usr"
        result = extract_identifiers(text)
        # @example might match (4+ chars after @), but @usr (3 chars) should not
        assert not any("@usr" == r for r in result)

    def test_username_min_length(self):
        # @ab is too short (3 chars total, need 5+), @abcde is valid
        text = "@ab @abcde"
        result = extract_identifiers(text)
        assert "@ab" not in result
        assert "@abcde" in result

    def test_invite_links(self):
        text = "Join https://t.me/+AbCdEf123 for the group"
        result = extract_identifiers(text)
        assert "https://t.me/+AbCdEf123" in result

    def test_empty_text(self):
        assert extract_identifiers("") == []

    def test_no_identifiers(self):
        assert extract_identifiers("just some random text without any links") == []

    def test_does_not_match_not_tme(self):
        text = "Visit not.me/channel for something else"
        result = extract_identifiers(text)
        # Should not match because of negative lookbehind
        assert not any("not.me" in r for r in result)

    def test_csv_like_text(self):
        text = '"@chan1","name","https://t.me/chan2"\n"@chan3","other","-1001234567890"'
        result = extract_identifiers(text)
        assert "@chan1" in result
        assert "https://t.me/chan2" in result
        assert "@chan3" in result
        assert "-1001234567890" in result


class TestParseFile:
    def _make_xlsx(self, data: list[list]) -> bytes:
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_txt_bytes(self):
        content = b"@channel1\nhttps://t.me/channel2\n-1001234567890"
        result = parse_file(content, "channels.txt")
        assert "@channel1" in result
        assert "https://t.me/channel2" in result
        assert "-1001234567890" in result

    def test_csv_bytes(self):
        content = b'"name","link"\n"Test","https://t.me/ch1"\n"Other","@chan2"'
        result = parse_file(content, "data.csv")
        assert "https://t.me/ch1" in result
        assert "@chan2" in result

    def test_csv_bom(self):
        content = b"\xef\xbb\xbf@channel1\n@channel2"
        result = parse_file(content, "bom.csv")
        assert "@channel1" in result
        assert "@channel2" in result

    def test_xlsx_bytes(self):
        data = [
            ["Name", "Link"],
            ["Chan1", "https://t.me/channel1"],
            ["Chan2", "@mychannel"],
        ]
        content = self._make_xlsx(data)
        result = parse_file(content, "channels.xlsx")
        assert "https://t.me/channel1" in result
        assert "@mychannel" in result

    def test_xlsx_negative_ids(self):
        data = [["ID"], ["-1001234567890"], ["-1009876543210"]]
        content = self._make_xlsx(data)
        result = parse_file(content, "ids.xlsx")
        assert "-1001234567890" in result
        assert "-1009876543210" in result

    def test_xlsx_multiple_sheets(self):
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.append(["@chan1"])
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["https://t.me/chan2"])
        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        result = parse_file(content)
        assert "@chan1" in result
        assert "https://t.me/chan2" in result

    def test_xlsx_empty(self):
        content = self._make_xlsx([])
        result = parse_file(content, "empty.xlsx")
        assert result == []

    def test_binary_garbage(self):
        content = bytes(range(256))
        result = parse_file(content, "garbage.bin")
        # Should not crash, may return empty or some matches from latin-1 decode
        assert isinstance(result, list)

    def test_no_filename(self):
        content = b"@testchan\nhttps://t.me/other"
        result = parse_file(content)
        assert "@testchan" in result
        assert "https://t.me/other" in result

    def test_latin1_fallback(self):
        # Invalid utf-8 but valid latin-1 with an identifier
        content = b"@validchan\xff\xfe some text"
        result = parse_file(content, "weird.txt")
        assert "@validchan" in result


class TestDeduplicateIdentifiers:
    def test_case_insensitive(self):
        identifiers = ["@Channel1", "@channel1", "@CHANNEL1"]
        result = deduplicate_identifiers(identifiers)
        assert result == ["@Channel1"]

    def test_preserves_order(self):
        identifiers = ["@ch3", "@ch1", "@ch2", "@ch1"]
        result = deduplicate_identifiers(identifiers)
        assert result == ["@ch3", "@ch1", "@ch2"]

    def test_strips_whitespace(self):
        identifiers = ["@ch1 ", " @ch1", "@ch1"]
        result = deduplicate_identifiers(identifiers)
        assert result == ["@ch1 "]

    def test_empty_list(self):
        assert deduplicate_identifiers([]) == []

    def test_no_duplicates(self):
        identifiers = ["@ch1", "@ch2", "@ch3"]
        result = deduplicate_identifiers(identifiers)
        assert result == identifiers


class TestNormalizeIdentifier:
    """Tests for normalize_identifier function."""

    def test_at_username(self):
        value, kind = normalize_identifier("@TestChannel")
        assert value == "testchannel"
        assert kind == "username"

    def test_at_username_upper(self):
        value, kind = normalize_identifier("@TESTCHANNEL")
        assert value == "testchannel"
        assert kind == "username"

    def test_tme_link_https(self):
        value, kind = normalize_identifier("https://t.me/channel1")
        assert value == "channel1"
        assert kind == "username"

    def test_tme_link_bare(self):
        value, kind = normalize_identifier("t.me/channel1")
        assert value == "channel1"
        assert kind == "username"

    def test_tme_link_with_post(self):
        value, kind = normalize_identifier("t.me/channel/123")
        assert value == "channel"
        assert kind == "username"

    def test_tme_link_https_with_post(self):
        value, kind = normalize_identifier("https://t.me/channel/456")
        assert value == "channel"
        assert kind == "username"

    def test_numeric_id_negative(self):
        value, kind = normalize_identifier("-1001234567890")
        assert value == "-1001234567890"
        assert kind == "numeric_id"

    def test_numeric_id_positive(self):
        value, kind = normalize_identifier("123456789")
        assert value == "123456789"
        assert kind == "numeric_id"

    def test_bare_username(self):
        value, kind = normalize_identifier("testchan")
        assert value == "testchan"
        assert kind == "username"

    def test_bare_username_short(self):
        # Too short (3 chars) doesn't match bare username pattern (needs 4+ chars)
        value, kind = normalize_identifier("abc")
        assert kind == "unknown"

    def test_empty_string(self):
        value, kind = normalize_identifier("")
        assert value == ""
        assert kind == "unknown"

    def test_whitespace_only(self):
        value, kind = normalize_identifier("   ")
        assert value == ""
        assert kind == "unknown"

    def test_unknown_identifier(self):
        value, kind = normalize_identifier("garbage!%^&*")
        assert value == "garbage!%^&*"
        assert kind == "unknown"

    def test_strips_whitespace(self):
        value, kind = normalize_identifier("  @TestChannel  ")
        assert value == "testchannel"
        assert kind == "username"
